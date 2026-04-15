"""
csv_parser.py — Parse mileage logs and expense sheets from CSV / Excel.

Handles:
  - GPS module mileage export (from GENESIS/modules/gps)
  - Manual mileage log CSVs (date, start/end odometer, miles, purpose)
  - Expense sheets (date, vendor, amount, category)
  - Excel workbooks (.xlsx) via openpyxl
  - Comma or tab-separated values

Output: dicts compatible with load_accountant_data()
"""

import csv
import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Union

logger = logging.getLogger("irs_forms_export.parsers.csv")

# ── Column name aliases ────────────────────────────────────────────────────────

_MILES_ALIASES = {
    "miles", "business_miles", "total_miles", "distance", "trip_miles",
    "business miles", "total miles", "miles driven",
}
_DATE_ALIASES = {
    "date", "trip_date", "service_date", "transaction_date",
    "trip date", "service date", "log date",
}
_PURPOSE_ALIASES = {
    "purpose", "trip_purpose", "description", "notes", "category",
    "trip purpose", "type", "reason",
}
_AMOUNT_ALIASES = {
    "amount", "total", "cost", "charge", "price", "expense",
    "amount_usd", "total_usd", "amount_charged",
}
_VENDOR_ALIASES = {
    "vendor", "merchant", "store", "payee", "business",
    "vendor_name", "merchant_name", "description",
}
_CATEGORY_ALIASES = {
    "category", "type", "expense_type", "account",
}


def _normalize_header(h: str) -> str:
    return h.lower().strip().replace(" ", "_").replace("-", "_")


def _find_col(headers: list[str], aliases: set) -> str | None:
    """Find which header matches a set of aliases."""
    normalized = {_normalize_header(h): h for h in headers}
    for alias in aliases:
        alias_n = _normalize_header(alias)
        if alias_n in normalized:
            return normalized[alias_n]
    return None


def _parse_dollar(val: str) -> float:
    """Parse a dollar string like '$1,234.56' to float."""
    if not val:
        return 0.0
    cleaned = str(val).replace("$", "").replace(",", "").strip()
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return 0.0


def _parse_float(val: str) -> float:
    if not val:
        return 0.0
    try:
        return float(str(val).replace(",", "").strip())
    except ValueError:
        return 0.0


def _parse_date(val: str) -> str:
    """Normalize a date string to YYYY-MM-DD."""
    if not val:
        return ""
    val = str(val).strip()
    # Already ISO
    if len(val) == 10 and val[4] == "-":
        return val
    # Try common formats
    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%B %d, %Y",
                "%b %d, %Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return val  # Return as-is if parsing fails


def _is_business_purpose(purpose: str) -> bool:
    """Return True if the trip purpose is business (not personal/commute)."""
    if not purpose:
        return True  # Default to business for mileage logs
    p = purpose.lower()
    personal_keywords = {"personal", "commute", "home", "vacation", "medical",
                         "charity", "school", "errand"}
    business_keywords = {"delivery", "amazon", "flex", "business", "client",
                         "meeting", "work", "pickup", "dropoff"}
    if any(k in p for k in business_keywords):
        return True
    if any(k in p for k in personal_keywords):
        return False
    return True  # Assume business if unclear


def _read_csv_rows(source: Union[str, Path, bytes, io.StringIO]) -> tuple[list, list]:
    """
    Read CSV/TSV and return (headers, rows).
    Handles bytes, file path, or StringIO.
    """
    if isinstance(source, bytes):
        # Try UTF-8 first, then latin-1
        try:
            text = source.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = source.decode("latin-1")
        f = io.StringIO(text)
    elif isinstance(source, (str, Path)):
        f = open(str(source), "r", encoding="utf-8-sig", errors="replace")
    else:
        f = source

    # Sniff delimiter
    sample = f.read(4096)
    f.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(f, dialect=dialect)
    headers = reader.fieldnames or []
    rows = list(reader)

    if hasattr(f, "close") and not isinstance(source, io.StringIO):
        f.close()

    return list(headers), rows


def _read_excel_rows(source: Union[str, Path, bytes],
                     sheet_name: str = None) -> tuple[list, list]:
    """Read Excel file using openpyxl. Returns (headers, rows as dicts)."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("csv_parser: openpyxl not installed — run: pip install openpyxl")

    if isinstance(source, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
    else:
        wb = openpyxl.load_workbook(str(source), data_only=True)

    ws = wb[sheet_name] if sheet_name else wb.active

    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return [], []

    headers = [str(h or "").strip() for h in rows_raw[0]]
    rows = []
    for raw_row in rows_raw[1:]:
        if any(v is not None for v in raw_row):
            rows.append({headers[i]: raw_row[i] for i in range(len(headers))})

    return headers, rows


# ── Mileage log parser ────────────────────────────────────────────────────────

def parse_mileage_csv(source: Union[str, Path, bytes],
                      tax_year: int = 2024,
                      irs_rate: float = None) -> dict:
    """
    Parse a mileage log CSV/Excel file.

    Expected columns (flexible naming):
        date, miles (or distance), purpose (optional)
        OR: date, odometer_start, odometer_end, purpose

    Returns dict compatible with load_accountant_data() mileage key:
        {
            "total_business_miles": float,
            "irs_rate": float,
            "trips": [{"date": ..., "miles": ..., "purpose": ...}, ...],
            "_source": "csv",
            "_rows_parsed": int,
            "_rows_skipped": int,
        }
    """
    from ..validator import YEAR_RULES
    if irs_rate is None:
        rules = YEAR_RULES.get(tax_year, YEAR_RULES[2024])
        irs_rate = rules["mileage_rate"]

    # Detect Excel vs CSV
    is_excel = False
    if isinstance(source, (str, Path)):
        is_excel = Path(source).suffix.lower() in (".xlsx", ".xls", ".xlsm")
    elif isinstance(source, bytes):
        is_excel = source[:4] in (b"PK\x03\x04",)  # ZIP signature = xlsx

    if is_excel:
        headers, rows = _read_excel_rows(source)
    else:
        headers, rows = _read_csv_rows(source)

    # Identify columns
    col_date    = _find_col(headers, _DATE_ALIASES)
    col_miles   = _find_col(headers, _MILES_ALIASES)
    col_purpose = _find_col(headers, _PURPOSE_ALIASES)
    col_odo_start = _find_col(headers, {"odometer_start", "start_odometer",
                                         "odo_start", "start_miles"})
    col_odo_end   = _find_col(headers, {"odometer_end", "end_odometer",
                                         "odo_end", "end_miles"})

    total_business_miles = 0.0
    trips = []
    skipped = 0

    for row in rows:
        # Get miles from direct column or odometer diff
        miles = 0.0
        if col_miles:
            miles = _parse_float(str(row.get(col_miles, 0) or 0))
        elif col_odo_start and col_odo_end:
            start = _parse_float(str(row.get(col_odo_start, 0) or 0))
            end   = _parse_float(str(row.get(col_odo_end, 0) or 0))
            miles = max(end - start, 0.0)

        if miles <= 0:
            skipped += 1
            continue

        # Get purpose
        purpose = str(row.get(col_purpose, "") or "").strip() if col_purpose else ""
        if not _is_business_purpose(purpose):
            skipped += 1
            continue

        # Get date
        date = ""
        if col_date:
            raw_date = row.get(col_date, "")
            if raw_date:
                date = _parse_date(str(raw_date))

        total_business_miles += miles
        trips.append({
            "date": date,
            "miles": round(miles, 1),
            "purpose": purpose or "Business",
        })

    return {
        "total_business_miles": round(total_business_miles, 1),
        "irs_rate": irs_rate,
        "computed_deduction": round(total_business_miles * irs_rate, 2),
        "trips": trips,
        "_source": "csv",
        "_rows_parsed": len(trips),
        "_rows_skipped": skipped,
    }


# ── Expense sheet parser ──────────────────────────────────────────────────────

_CATEGORY_MAP = {
    # Keywords → Schedule C line
    "fuel":       "car_and_truck",
    "gas":        "car_and_truck",
    "vehicle":    "car_and_truck",
    "supplies":   "supplies",
    "supply":     "supplies",
    "amazon":     "supplies",
    "phone":      "phone_internet",
    "internet":   "phone_internet",
    "cell":       "phone_internet",
    "data":       "phone_internet",
    "software":   "software_subscriptions",
    "app":        "software_subscriptions",
    "travel":     "travel",
    "hotel":      "travel",
    "flight":     "travel",
    "meals":      "meals",
    "food":       "meals",
    "restaurant": "meals",
    "repairs":    "repairs",
    "maintenance": "repairs",
    "insurance":  "insurance",
    "legal":      "legal_professional",
    "advertising": "advertising",
    "marketing":  "advertising",
    "office":     "office",
    "education":  "education_training",
    "training":   "education_training",
    "rent":       "rent_other_property",
    "utilities":  "utilities",
}


def _map_category(raw: str) -> str:
    """Map a raw category string to a ScheduleCExpenses field name."""
    if not raw:
        return "other_expenses"
    raw_lower = raw.lower()
    for keyword, field in _CATEGORY_MAP.items():
        if keyword in raw_lower:
            return field
    return "other_expenses"


def parse_expense_csv(source: Union[str, Path, bytes],
                      tax_year: int = 2024) -> dict:
    """
    Parse an expense sheet CSV/Excel file.

    Expected columns: date, vendor/description, amount, category (optional)

    Returns dict compatible with load_accountant_data() self_employment_income
    expenses sub-dict, plus a raw list of transactions.
        {
            "expenses": {ScheduleCExpenses fields...},
            "transactions": [{date, vendor, amount, category}, ...],
            "_source": "csv",
            "_rows_parsed": int,
        }
    """
    is_excel = False
    if isinstance(source, (str, Path)):
        is_excel = Path(source).suffix.lower() in (".xlsx", ".xls", ".xlsm")
    elif isinstance(source, bytes):
        is_excel = source[:4] in (b"PK\x03\x04",)

    if is_excel:
        headers, rows = _read_excel_rows(source)
    else:
        headers, rows = _read_csv_rows(source)

    col_date     = _find_col(headers, _DATE_ALIASES)
    col_amount   = _find_col(headers, _AMOUNT_ALIASES)
    col_vendor   = _find_col(headers, _VENDOR_ALIASES)
    col_category = _find_col(headers, _CATEGORY_ALIASES)

    # Accumulate into ScheduleCExpenses field names
    totals = {
        "advertising": 0.0, "car_and_truck": 0.0, "supplies": 0.0,
        "phone_internet": 0.0, "software_subscriptions": 0.0,
        "travel": 0.0, "meals": 0.0, "repairs": 0.0, "insurance": 0.0,
        "legal_professional": 0.0, "office": 0.0, "education_training": 0.0,
        "rent_other_property": 0.0, "utilities": 0.0, "other_expenses": 0.0,
    }
    transactions = []

    for row in rows:
        amount = _parse_dollar(str(row.get(col_amount, 0) or 0)) if col_amount else 0.0
        if amount <= 0:
            continue

        vendor = str(row.get(col_vendor, "") or "").strip() if col_vendor else ""
        raw_cat = str(row.get(col_category, "") or "").strip() if col_category else ""
        date = _parse_date(str(row.get(col_date, "") or "")) if col_date else ""

        # Determine category — prefer explicit category column, then vendor keywords
        if raw_cat:
            sched_c_field = _map_category(raw_cat)
        else:
            sched_c_field = _map_category(vendor)

        totals[sched_c_field] = totals.get(sched_c_field, 0.0) + amount
        transactions.append({
            "date": date,
            "vendor": vendor,
            "amount": amount,
            "category": sched_c_field,
            "raw_category": raw_cat,
        })

    return {
        "expenses": {k: round(v, 2) for k, v in totals.items()},
        "transactions": transactions,
        "_source": "csv",
        "_rows_parsed": len(transactions),
    }
